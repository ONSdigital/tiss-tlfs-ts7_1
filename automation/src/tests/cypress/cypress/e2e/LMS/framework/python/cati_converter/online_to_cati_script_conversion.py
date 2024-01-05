# function to remove empty rows
import openpyxl
import os
import pathlib
import random
import functions as func

# hours = ['UPOT1', 'UPOT', 'UPOT2', 'CASAC', 'USHR', 'USHR1', 'USHR2', 'ACTHR', 'ACTHR1', 'ACTHR2',  'POT', 'POT1', 'POT2']
removals = ['COUNTRY_QUAL_INTRO', 'ALTADD', 'ALTADDTYPE', 'TIMEAWAY']
# pdwg = ['PDWG20', 'PDWG21', 'PDWG22']
colouriser = ['GRSSKP', 'GRSSKP1A', 'GRSSKP1B', 'GRSSKP2', 'GRSSKP2A', 'GRSSKP2B', 'GRSSKPA']

from openpyxl.styles import PatternFill

def convert_online_spreadsheets_to_cati(repo_name='', input_path='/app/data/LMS_Spreadsheets',
                                        output_path = "/app/data/LMS_Spreadsheets_completed"):
    new_path = pathlib.Path().resolve()
    new_path = str(new_path)
    folders = new_path.split("/")

    #These 2 will need changing to run elsewhere
    # path_to_lms = '/tmp/build/' + folders[3] + '/' + repo_name + '/automation/src/tests/cypress/cypress/e2e/LMS/'
    # path_to_lms_live = '/tmp/build/' + folders[3] + '/' + repo_name + '/automation/src/tests/cypress/cypress/e2e/LMS/'
    answers_file = '/tmp/build/' + folders[3] + '/' + repo_name + '/automation/src/tests/cypress/cypress/e2e/LMS/framework/command_answer_changes.tsv'
    # input_path = '/app/data/LMS_Spreadsheets'
    # output_path = "/app/data/LMS_Spreadsheets_completed"
    # answers_file = '/Users/ianfoulsham/GitHub/TISS-CTF-LMS-TS5/automation/src/tests/cypress/cypress/e2e/LMS/framework/command_answer_changes.tsv'
    # completed_path = '/Users/ianfoulsham/TO_Misc'
    # required_survey = 'lms'
    # required_country = 'england'

    # completed_files = []
    with open(answers_file) as af:
        answers_list = af.readlines()
        answers_list = [x.strip() for x in answers_list]

    ds_store = input_path + '/.DS_Store'
    os.system('rm "{}"'.format(ds_store))
    # print("one")
    for dirpath, dirnames, files in os.walk(input_path):
        print(f'Found directory: {dirpath}')
        for my_file_name in files:
            print(my_file_name)
            # load excel file
            book = openpyxl.load_workbook(input_path + '/' + my_file_name)
            # select the sheet
            sheet = book['Paste']
            #
            people_and_names = func.get_person_names(sheet)
            print("Person list is {}".format(people_and_names))
            # Need to put the person in here maybe?  Or have a function for ECAUTH
            func.change_answers(sheet, answers_list, people_and_names)

            # Add "No" to the Employment_Intro
            list_emp_intro = func.find_command(sheet, "PAYINTRO")
            print(list_emp_intro)
            print(len(list_emp_intro))
            if list_emp_intro != "":
                list_emp_intro_occurrences = list_emp_intro.split(",")
                for value in list_emp_intro_occurrences:
                    func.set_answer(sheet, value, "No")
                    func.change_command(sheet, value, "PAYINTRO", "PAYINTRO_TO")

            output_file_name = my_file_name.replace("online", "to")
            book.save(output_path + '/' + output_file_name )
            book = openpyxl.load_workbook(output_path + '/' + output_file_name)
            sheet = book['Paste']
            sheet = func.colourise(sheet, colouriser, "FFA500")
            print("rows before removing:", sheet.max_row)


            #
            #Add the rows at the start of all the TO scripts
            #
            sheet = func.add_telephone_start_rows(sheet)
            book.save(output_path + '/' + output_file_name )
            book = openpyxl.load_workbook(output_path + '/' + output_file_name)
            sheet = book['Paste']
            #Doing the adding of INTUSE and Employment_Intro in 2 parts.
            #Part 1 is finding all locations of the point at which lines are inserted (above)
            paid_job_found = func.find_command(sheet, "PAIDJOB")
            # print(paid_job_found)

            if paid_job_found != "":
                sheet = func.add_paid_job_rows(sheet, my_file_name, paid_job_found)

            book.save(output_path + '/' + output_file_name )
            book = openpyxl.load_workbook(output_path + '/' + output_file_name)
            sheet = book['Paste']

            address_hist_found = func.find_command(sheet, "ADDRHIST_CHECK")
            # print(address_hist_found)
            if address_hist_found != "":
                sheet = func.add_addr_intro_rows(sheet, my_file_name, address_hist_found)
            #colorise(sheet)
            # Removed these as they were to sort out the passport section generically and can't really be run twice
            # add_passport_rows(sheet,1)
            # add_passport_rows(sheet,2)

            book.save(output_path + '/' + output_file_name)
            book = openpyxl.load_workbook(output_path + '/' + output_file_name)
            sheet = book['Paste']
            #Add AddrHist_Intro before AddrHist_Check
            book.save(output_path + '/' + output_file_name)
            #Start of the complicated AltAddCheck moving from the individual to Household


            #Get the answers that each person gives
            list_of_altaddr_people = func.get_altaddr_answers(sheet)
            print("exitted with alt address answers {}".format(list_of_altaddr_people))
            answer_for_altadd = []
            if len(list_of_altaddr_people) != 0:
                print("Going in to the list of alt address people")
                for items in people_and_names:

                    if items[0] in list_of_altaddr_people[0]:
                        print("matching name for person who answered yes: {}".format(items[1]))
                        answer_for_altadd.append(items[1])

                # print(answer_for_altadd)
                # AltAddAnswer
            print("answer for alt address is {}".format(answer_for_altadd))
            if len(answer_for_altadd) != 0:
                alt_addr_answer = ";".join(answer_for_altadd)
            else:
                alt_addr_answer = "None of these people"

                # print("alt_addr_answer = {}".format(alt_addr_answer))

            altaddr_row = func.find_commands_by_priority(sheet, ["TIED", "TEN1SomeoneElse_OTH", "TEN1SomeoneElse",
                                                                    "Ten1"])
            print("Alt_addrRow is {}".format(altaddr_row))
            print("Processing AltAddressChk")
            if len(altaddr_row) > 0:
                sheet = func.add_row(sheet, int(altaddr_row) + 1, "AltAddressChk", alt_addr_answer)
            book.save(output_path + '/' + output_file_name)
            book = openpyxl.load_workbook(output_path + '/' + output_file_name)
            sheet = book['Paste']

            if len(list_of_altaddr_people) != 0:
                #At this point we have the actual values to be used in a list of lists.
                #Plus we have the location of the new AltAddressChk row.
                #So what we need to do is remove all the rows in the individual sections and then use the list


                book.save(output_path + '/' + output_file_name )
                #to populate the rows afterwards - including "Person"
                #might there be a problem with setting the person back at the end?
                #Probs need to understand what the person is before anything happens - we might have that already though
                alt_addr_new_pos = func.find_command(sheet, "AltAddressChk")
                # add_row(sheet, int(altaddr_row) + 1, "AltAddressChk", alt_addr_answer)
                book.save(output_path + '/' + output_file_name)

                print("rows after processing:",sheet.max_row)
                for list_item in list_of_altaddr_people:
                    # print("List item is: {}".format(list_of_altaddr_people))
                    length = len(list_item)
                    sheet = func.add_row(sheet, int(alt_addr_new_pos) + 1, "Person", list_item[0])
                    book.save(output_path + '/' + output_file_name)
                    book = openpyxl.load_workbook(output_path + '/' + output_file_name)
                    sheet = book['Paste']

                    for i in range(1, length):
                        # print("Removing: {}".format(list_item[i][0]))

                        func.remove(sheet, list_item[i][0])
                        book.save(output_path + '/' + output_file_name)
                        if list_item[i][0] == "ALTADDTYPE":
                            list_item[i][0] = "OTHADDTYPE"
                        if list_item[i][0] == "TIMEAWAY":
                            list_item[i][0] = "TIMEAWAY_TO"
                        sheet = func.add_row(sheet, int(alt_addr_new_pos) + i + 1, list_item[i][0], list_item[i][1])
                        book.save(output_path + '/' + output_file_name)
                        book = openpyxl.load_workbook(output_path + '/' + output_file_name)
                        sheet = book['Paste']


            #Find and convert OWNBUS18
            ownbus_locations = func.find_command(sheet, "OWNBUS18")
            # print("ownbus_locations {}".format(ownbus_locations))
            #start off with getting one person to work
            if ownbus_locations != "":
                ownbus_individual_location = ownbus_locations.split(",")
                # print("ownbus_individual_location {}".format(ownbus_individual_location))
                #change the value to the TO equivalent
                for ownbus_position in ownbus_individual_location:

                    func.change_command(sheet, ownbus_position, "OWNBUS18", "OWNBUS18_Int1")
                    #get the answer
                    ownbus_answer = func.get_answer(sheet, ownbus_position)
                    if "YES" in ownbus_answer.upper():

                        func.set_answer(sheet, ownbus_position, "Yes")
                        #add a row after
                        int2_answer = ownbus_answer.replace("Yes, for ", "")
                        sheet = func.add_row(sheet, int(ownbus_position) + 1, "OWNBUS18_Int2", int2_answer )

                        book.save(output_path + '/' + output_file_name)
                book.save(output_path + '/' + output_file_name)
            book = openpyxl.load_workbook(output_path + '/' + output_file_name)
            sheet = book['Paste']
            #Find and convert OWNBUS18
            ownbus_catch_locations = func.find_command(sheet, "OWNBUS18_Catch")
            # print("ownbus_locations {}".format(ownbus_locations))
            #start off with getting one person to work
            if ownbus_catch_locations != "":
                ownbus_catch_ind_location = ownbus_catch_locations.split(",")
                # print("ownbus_individual_location {}".format(ownbus_individual_location))
                #change the value to the TO equivalent
                for ownbus_catch_position in ownbus_catch_ind_location:

                    func.change_command(sheet, ownbus_catch_position, "OWNBUS18_Catch", "OWNBUS18_Catch_Int1")
                    #get the answer
                    ownbus_catch_answer = func.get_answer(sheet, ownbus_catch_position)
                    if "YES" in ownbus_catch_answer.upper():

                        func.set_answer(sheet, ownbus_catch_position, "Yes")
                        #add a row after
                        catch_int2_answer = ownbus_catch_answer.replace("Yes, for ", "")
                        sheet = func.add_row(sheet, int(ownbus_catch_position) + 1, "OWNBUS18_Catch_Int2", catch_int2_answer )

                        book.save(output_path + '/' + output_file_name)
            book.save(output_path + '/' + output_file_name)
            book = openpyxl.load_workbook(output_path + '/' + output_file_name)
            sheet = book['Paste']
            for cmd in removals:
                sheet = func.remove(sheet, cmd)
                book.save(output_path + '/' + output_file_name )
                book = openpyxl.load_workbook(output_path + '/' + output_file_name)
                sheet = book['Paste']

            #Change the everot question to everot_int1 and infer the answer
            everot_locations = func.find_command(sheet, "EVEROT")
            if everot_locations != "":

                sheet = func.convert_everot(sheet, everot_locations)

            book.save(output_path + '/' + output_file_name)