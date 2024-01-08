from datetime import date

def colourise(col_sheet, cmd_to_colour, colour_to_add):
    from openpyxl.styles import PatternFill
    for row in col_sheet.iter_rows():
        for command in cmd_to_colour:
            if str(row[0].internal_value).upper() == command.upper():
                row[0].fill = PatternFill("solid", start_color=colour_to_add)
                row[1].fill = PatternFill("solid", start_color=colour_to_add)
                print('colorised row {}'.format(row[1].internal_value))
    return col_sheet

def convert_name(name_string):
    # print("Name to convert is {}".format(name_string))
    names = name_string.split(";")
    formatted_name = names[0].strip() + " " + names[1].strip() + " " + names[2].strip()
    reformatted_name = formatted_name.replace("  ", " ")
    print("Formatted name is {}".format(reformatted_name))
    return reformatted_name

# common function to calculate a person's current age in years
def calculate_age(born):
    today = date.today()
    return today.year - born.year - ((today.month, today.day) < (born.month, born.day))

def change_answers(my_sheet, answer_list, person_list):
    # print("answers")
    for row in my_sheet:
        for item in answer_list:
            command = item.split('\t')

            # print("command is {}".format(command[0].upper()))

            # Locate the command in the test spreadsheet that corresponds to a command in the list
            #  of values that might need to be changed. Not all values will need changing
            if str(command[0]).upper() == str(row[0].internal_value).upper():
                # print("Found {} in row".format(command[0].upper()))
                # print(command[1])
                # print(str(row[1].internal_value))

                #  This is the answer in column 2 of the test spreadsheet that will need checking for changes.
                # It could be more than one value, so each one will need checking
                answer_part = str(row[1].internal_value).split(";")

                # blank values in the spreadsheet were showing as Python's None value which got str'd to "None" which
                # killed the test.
                if row[1].internal_value is None:
                    answer_part = ""

                replacement_list = []
                for part in answer_part:
                    # part is the answer component in the test spreadsheet
                    # We only need to go through this part if there is a * in the command
                    # The point is to replace the * with the name from the name list and then check for
                    # a match between the amended command[1] text and part
                    if "*" in command[1]:
                        for person in person_list:
                            new_text = command[1].replace("*", person[1])
                            if new_text in str(part):
                                command[1] = new_text
                                break
                    if command[1] in str(part):
                        print("found a value to change in command {}: from {} to {}".format(command[0].upper(), command[1], str(command[2])))
                        replacement_list.append(command[2])
                    else:
                        replacement_list.append(str(part))

                row[1].value = ";".join(replacement_list)

    return my_sheet



def check_for_commands(my_sheet, command_list):
    # from openpyxl.styles import PatternFill
    value_found = False
    for row in my_sheet.iter_rows():
        for command in command_list:
            if str(row[0].internal_value).upper() == command.upper():
                value_found = True

    return value_found

def add_to_start_rows(my_sheet):
    #rows = list(sheet.iter_rows(min_row=1, max_row=sheet.max_row))
    my_sheet.insert_rows(1,2)
    my_sheet['A1'] = "AddChk1"
    my_sheet['B1'] = "Yes"
    my_sheet['A2'] = "Household_Intro"
    print("Telephone Start Rows added")
    return my_sheet

def find_commands_by_priority(my_sheet, commands):

    #The point of this method is to find a row, usually to subsequently insert a new command
    #where the command can come after a number of possible previous commands.
    #For instance, when converting to TO from online, AltAddressChk could be after Ten1, Ten1SomeoneElse,
    #Ten1SomeoneElse_Oth or TIED. TIED is the highest priority and Ten1 the lowest, so we go through in
    #priority order until we find the place to insert a row with AltAddressChk.
    # NB - I don't think this will work at the individual level where there may be repeats of the same commands for different people

    command_occurrence = ""
    for command in commands:
        counter = 0
        for row in my_sheet.iter_rows():
            counter += 1
            if str(row[0].internal_value).upper() == command.upper():
                command_occurrence = str(counter)
                print("value found for altaddress check is {}".format(str(row[0].internal_value).upper()))
                return command_occurrence
    #If nothing is found, we return the blank initial value
    return command_occurrence

def add_addr_intro_rows(my_sheet, my_file, appearances):
    rows_to_add = 1
    #where there are more than 1 item in the list of places for paidjob, each time rows are added, the next place
    #to insert rows is no longer correct, so the uplift gets 1 or 2 added to it, depending on whether it is lms or lmb
    uplift = 0
    insert_row_locn = appearances.split(",")
    # print(len(insert_row_locn))
    for item in insert_row_locn:
        uplift = rows_to_add * insert_row_locn.index(item)
        # print("uplift = {}".format(uplift) )
        location = int(item) + uplift
        # print("location is: {}".format(location))
        # print(rows_to_add)
        my_sheet.insert_rows(location,rows_to_add)

        my_sheet['A{}'.format(location)] = "AddrHist_Intro"

    return my_sheet


def get_person_age(my_sheet):
    print("get age")
    retained_person = ""
    age_list = []
    for row in my_sheet.iter_rows():
        # This sets the person value for the list
        if str(row[0].internal_value).upper() == "PERSON":
            retained_person = str(row[1].internal_value)
        # Where DOB is left blank the person will put in their age, so no calulation required
        if str(row[0].internal_value).upper() == "AGE":
            returned_age = str(row[1].internal_value)
            age_list.append([retained_person, returned_age])
        # Where the DOB is entered, this gets more complex -
        # have to create the date and then work out the age from today
        if str(row[0].internal_value).upper() == "DOB":
            if row[1].internal_value != "" and row[1].internal_value != None :
                import datetime

                date_of_birth = row[1].internal_value.split(";")
                if date_of_birth[2] != '':
                    # print("dob = {}".format(date_of_birth) )
                    birth_date = datetime.datetime(int(date_of_birth[2]), int(date_of_birth[1]), int(date_of_birth[0]))
                    # print("DOB i {}".format(date_of_birth))
                    # print("Birth date is {}".format(birth_date))
                    my_age = calculate_age(birth_date)
                    age_list.append([retained_person, my_age])

    # print("age list is: {}".format(age_list))
    return age_list

def add_new_job_rows(workbook_path, commands, new_row_values, position = "Before"):
    counter = 0
    import openpyxl
    for command in commands:
        book = openpyxl.load_workbook(workbook_path)
        my_sheet = book['Paste']
        new_row_value = new_row_values[counter]
        counter += 1
        new_locations = find_command(my_sheet, command)
        if new_locations:
            # print("New Location is {}".format(new_locations))
            location = new_locations.split(",")
            for row in location:
                new_value = get_answer(my_sheet, row)
                add_row(my_sheet, row, new_row_value, new_value, position)
                book.save(workbook_path)
    # return my_sheet

def get_answer(my_sheet, index_location):

    return my_sheet[index_location][1].internal_value

def get_altaddr_answers(my_sheet):
    #first off we want to see if they have a yes in atladdr and if so, put the names in there
    # print("get_altaddr_answers")
    alt_addr_list = []
    alt_add_answers = []
    for row in my_sheet.iter_rows():

        if str(row[0].internal_value).upper() == "ALTADD" and str(row[1].internal_value).upper() == "YES":
            alt_addr_list.append([retained_person])

        if str(row[0].internal_value).upper() == "PERSON":
            retained_person = str(row[1].internal_value)

        if str(row[0].internal_value).upper() == "ALTADDTYPE":
            # print("at altaddtype")
            #go round the
            for i in range(len(alt_addr_list)) :
                # print("i={}".format(alt_addr_list[i]))
                for j in range(len(alt_addr_list[i])) :
                    # print("alt_addr_list[i][j]={}".format(alt_addr_list[i][j]))
                    if str(alt_addr_list[i][j]) == str(retained_person):
                        # print("Here we are {}".format(j))
                        alt_addr_list[i].insert(1, ["ALTADDTYPE", str(row[1].internal_value)])
                        # print("alt_addr_list after update with AltAddType: {}".format(alt_addr_list))
        if str(row[0].internal_value).upper() == "TIMEAWAY":
            for i in range(len(alt_addr_list)):
                # print("i={}".format(alt_addr_list[i]))

                # print("alt_addr_list[i][0]={}".format(alt_addr_list[i][j]))
                if str(alt_addr_list[i][0]) == str(retained_person):
                    # print("Here we are {}".format(0))
                    alt_addr_list[i].insert(2, ["TIMEAWAY", str(row[1].internal_value)])
                    # print("alt_addr_list after update with AltAddType: {}".format(alt_addr_list))
            # print("bar")

    # print("List of people with an alternate address: {}".format(alt_addr_list))
    return alt_addr_list

def get_person_names(my_sheet):
    # print("getting people")
    retained_person = ""
    name_list = []
    for row in my_sheet.iter_rows():

        if str(row[0].internal_value).upper() == "ABOUTYOU":
            returned_name = convert_name(str(row[1].internal_value))
            name_list.append(["1", returned_name])
        if str(row[0].internal_value).upper() == "PERSON":
            retained_person = str(row[1].internal_value)
        if str(row[0].internal_value).upper() in ["ADDPPL", "ADDSTU"]:
            returned_name = convert_name(str(row[1].internal_value))
            name_list.append([retained_person, returned_name])
        if str(row[0].internal_value).upper() == "SETPROXY":
            for pers in name_list:
                if retained_person == pers[0]:
                    indexy = name_list.index(pers)
                    name_list[indexy].append(str(row[1].internal_value))

    print("Full list of people: {}".format(name_list))

    return name_list


def convert_everot(my_sheet, everot_index):
    # print("everot_index at start: {}".format(everot_index))
    int2_answer = ""
    int1_answer = ""

    if everot_index == "":
        print("No EVEROT rows to convert")
        return my_sheet

    # else get the value of the answer in each everot occurrence
    full_index = everot_index.split(",")
    for value in reversed(full_index):
        # print("everot value is: {}".format(value))
        everot_answer = my_sheet[value][1].internal_value
        # print("answer to everot question is: {}".format(everot_answer))
        if "not work overtime" in everot_answer:
            int1_answer = "No"
        elif "both" in everot_answer:
            int1_answer = "Yes"
            int2_answer = "both paid and unpaid overtime"
        elif "unpaid" in everot_answer:
            int1_answer = "Yes"
            int2_answer = "unpaid overtime"
        elif "paid" in everot_answer:
            int1_answer = "Yes"
            int2_answer = "paid overtime"
        elif "self-employed" in everot_answer:
            int1_answer = "Self-employed and this doesn't apply"

        my_sheet[value][0].value = "EVEROT_INT1"
        my_sheet[value][1].value = int1_answer
        if int2_answer != "":
            my_sheet = add_row(my_sheet, int(value) + 1, "EVEROT_INT2", int2_answer)
        print("EVEROT rows are processed")

    return my_sheet


def set_answer(sheet, row_to_change, new_answer):

    sheet[row_to_change][1].value = new_answer
    return


def find_command(my_sheet, cmd_to_find):

    counter = 0
    already_found = "No"
    command_occurrences = ""

    for row in my_sheet.iter_rows():
        counter += 1

        if str(row[0].internal_value).upper() == cmd_to_find.upper():
            if already_found == "No":
                already_found = "Yes"
                command_occurrences = str(counter)
            else:
                command_occurrences = command_occurrences + "," + str(counter)
            # print("{} can be found in rows: {}".format(cmd_to_find, command_occurrences))
    return command_occurrences

def merge_list_by_person(list1, list2):
    print("merging")
    for first_entry in list1:
        for second_entry in list2:
            if first_entry[0] == second_entry[0]:
                first_entry.append(second_entry[1])
    return list1

def change_command(sheet, index_location, old_value, new_value):
    if sheet[index_location][0].internal_value.upper() == old_value.upper():
        # print("Changing {} to {} at row {}".format(old_value, new_value, index_location))
        sheet[index_location][0].value = new_value
        return 0
    else:
        return 1

def remove_sex(my_sheet, index_locations):
    index_location = index_locations.split(",")
    # print(len(index_location))
    sex = []
    for locn in index_location:
        if locn != '':
            # print(my_sheet[locn][0].internal_value)
            newval = my_sheet[locn][1].internal_value.split(";")
            my_sheet[locn][1].value = ';'.join(newval[:-1])
            sex.append([locn, newval[-1]])
            # print(sex)
    # print(sex)
    return sex



def get_command_and_person(sheet, cmd_to_find):

    counter = 0
    # already_found = "No"
    command_occurrences = []

    retained_person = ""

    for row in sheet.iter_rows():
        counter += 1
        if str(row[0].internal_value).upper() == "PERSON":
            retained_person = str(row[1].internal_value)
        if str(row[0].internal_value).upper() == cmd_to_find.upper():
            command_occurrences.append([retained_person, str(counter)])

            # print("{} can be found in rows: {}".format(cmd_to_find, command_occurrences))
    return command_occurrences

def add_row(my_sheet, appearances, command, value="", insert_point = "Before"):

    rows_to_add = 1
    uplift = 0
    extra_uplift = 0
    insert_row_locn = str(appearances).split(",")
    print(insert_row_locn)
    for item in insert_row_locn:
        if insert_point == "After":
            print("uplifting by 1")
            extra_uplift = 1
        else:
            print("No uplift")
            extra_uplift = 0
        print("insert_row_locn.index(item) is {}".format(insert_row_locn.index(item)))
        uplift = (rows_to_add * insert_row_locn.index(item)) + extra_uplift
        print("uplift = {}".format(uplift) )
        location = int(item) + uplift
        print("location is: {}".format(location))
        print(rows_to_add)
        my_sheet.insert_rows(int(location),int(rows_to_add))
        my_sheet['A{}'.format(location)] = command

        if value != "":
            my_sheet['B{}'.format(location)] = value

    return my_sheet

def remove(my_sheet, removable_commands):
    # iterate the sheet by rows
    # print("Removing {}".format(removable_commands))

    for row in my_sheet.iter_rows():
        if str(row[0].internal_value).upper() in removable_commands.upper():
            # delete the row
            # print("row to be deleted found {}".format(row[0].internal_value))
            # print(row[0].internal_value)
            my_sheet.delete_rows(row[0].row, 1)
            # recursively call the remove() with modified sheet data
            # remove(my_sheet, removable_commands)

    return my_sheet


def add_paid_job_rows(my_sheet, my_file_name, appearances):

    import random

    rows_to_add = 1
    # counting = 0
    #where there are more than 1 item in the list of places for paidjob, each time rows are added, the next place
    #to insert rows is no longer correct, so the uplift gets 1 or 2 added to it, depending on whether it is lms or lmb
    uplift = 0
    if "lmb" in my_file_name.lower():
        rows_to_add = 2
    insert_row_locn = appearances.split(",")
    # print(insert_row_locn)
    # print(len(insert_row_locn))
    for item in insert_row_locn:
        # print(item)
        # counting +=1
        #The list seems to be exercised twice, so this is a quick and dirty hack to stop that happening
        # if counting >= len(insert_row_locn):
        #     print("hit the breaks")
        #     break
        uplift = rows_to_add * insert_row_locn.index(item)
        # print("uplift = {}".format(uplift) )
        location = int(item) + uplift
        # print("location is: {}".format(location))
        # print(rows_to_add)
        my_sheet.insert_rows(location,rows_to_add)
        if rows_to_add == 2:
            # print(location)
            # print("Writing INTUSE at line ".format(location))
            my_sheet['A{}'.format(location)] = "INTUSE"
            intuse_random = random.choice(["Within the last three months", "Never used it?",
                                           "More than one year ago", "Between three months and a year ago," ])
            my_sheet['B{}'.format(location)] = intuse_random
            location += 1
            # print("Writing emp intro at line ".format(location))
            my_sheet['A{}'.format(location)] = "Employment_Intro"
        else:
            my_sheet['A{}'.format(location)] = "Employment_Intro"

    return my_sheet

def get_files_in_directory(my_path):
    #
    from os import listdir

    completed_directories = listdir(my_path)
    do_not_use = ['.DS_Store', 'Archived']
    full_list_of_files = []
    for dir in completed_directories:
        if dir not in do_not_use:
            all_files = listdir(my_path + '/' + dir)
            for file in all_files:
                new_file = file.replace('_to.xlsx', '_online.xlsx')
                # print(new_file)
                full_list_of_files.append(new_file)

    # print('These are the files to compare against {}'.format(full_list_of_files))
    return full_list_of_files



def add_telephone_start_rows(sheet):
    #rows = list(sheet.iter_rows(min_row=1, max_row=sheet.max_row))
    sheet.insert_rows(1,2)
    sheet['A1'] = "AddChk1"
    sheet['B1'] = "Yes"
    sheet['A2'] = "Household_Intro"